import pandas as pd
from bs4 import Tag

import openpyxl
from bs4 import BeautifulSoup
import requests
import csv

filename = "IPL1.csv"
f = open(filename, "w")
headers = "PLAYER NAME,TEAM,ROLE,MATCHES,NOT OUTs,RUNS,HIGHEST RUN,BATTING AVG,BALLS FACED,BATTING S.R,100s,50s,4s,6s,CATCHES,BALLS BOWLED,RUNS CONCEEDED,WICKETS,BEST BOWLING,BOWLING AVG,ECONOMY,BOWLING S.R\n"
f.write(headers)



class Player:

    def __init__(self,name:str,team:str, table:Tag):
        self.name=name
        self.team=team
        batcols=table[0].find_all('td')
        batcols = [ele.text.strip() for ele in batcols]
        self.totalmatches=batcols[1]
        self.totalnotouts=batcols[2]
        self.totalruns=batcols[3]
        self.highest = batcols[4]
        self.battingaverage = batcols[5]
        self.battingaverage=self.battingaverage.replace("-","0.0")
        if  float(self.battingaverage)<20:
            self.role="Bowler"
        else:
            self.role="Batsman"
        self.totalballsfaced = batcols[6]
        self.totalballsfaced=self.totalballsfaced.replace(",","")
        self.strikingrate = batcols[7]
        self.totalhundreds = batcols[8]
        self.totalfifties = batcols[9]
        self.totalfours = batcols[10]
        self.totalsixes = batcols[11]
        self.totalcatches=batcols[12]
        bwcols = table[1].find_all('td')
        bwcols = [ele.text.strip() for ele in bwcols]
        self.totalballsbowled= bwcols[2]
        self.totalballsbowled= self.totalballsbowled.replace(",","")
        self.totalrunsconceeded=bwcols[3]
        self.totalrunsconceeded=self.totalrunsconceeded.replace(",","")
        self.totalwickets = bwcols[4]
        self.bestbowlingfigure = bwcols[5]
        self.bowlingaverage = bwcols[6]
        self.economy = bwcols[7]
        self.bowlingstrikerate = bwcols[8]
        f.write(self.name+","+self.team+","+self.role+","+self.totalmatches+","+self.totalnotouts+","+self.totalruns+","+self.highest+","+self.battingaverage+","+self.totalballsfaced+","+self.strikingrate+","+self.totalhundreds+","+self.totalfifties+","+self.totalfours+","+self.totalsixes+","+self.totalcatches+","+self.totalballsbowled+","+self.totalrunsconceeded+","+self.totalwickets+","+self.bestbowlingfigure+","+self.bowlingaverage+","+self.economy+","+self.bowlingstrikerate+"\n")
        
def loadplayersfromexcel():
    wb = openpyxl.load_workbook('IPL.xlsx')
    ws = wb['Sheet1']
    i=1
    count=0
    cell = ws.cell(row=i, column=1)
    playerlist=[]
    while cell.value:
        if(cell.hyperlink):
            count=count+1
##            print(cell.value+" "+cell.hyperlink.target)
            playerlist.append(playerdetails(name=cell.value,url=cell.hyperlink.target))
        i=i+1
        cell=ws.cell(row=i,column=1)
    print(str(count))
    print(len(playerlist))
    #print(playerlist)
    # print(playerlist[3].name)

#Scrap from the given player url using beautiful soup
def playerdetails(name,url):
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    table = soup.find_all('tr', attrs={'class':'player-stats-table__highlight'})
    team=team_from_url(url) #punjab
    # print(name+' '+team)
    player=Player(table=table,name=name,team=team)
    return player

TEAMS=['chennai','bangalore','bengaluru','kolkata','rajasthan','mumbai','delhi','punjab','hyderabad']

def team_from_url(url):
    playerteam=''
    for team in TEAMS:
       if team in url:
           playerteam=team
    if playerteam=='bangalore':
        playerteam='bengaluru'
    return playerteam

loadplayersfromexcel()
f.close()







